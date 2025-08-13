from __future__ import annotations

from typing import Any, Dict, List
from io import BytesIO
from pathlib import Path
import tempfile
import subprocess
import shutil
import os
from datetime import datetime

from openpyxl import load_workbook
import streamlit as st

from core.blocks.base import BlockContext, ProcessingBlock
from core.errors import BlockException, BlockError, ErrorCode
from core.plan.logger import export_log


class ExcelReadDataBlock(ProcessingBlock):
    id = "excel.read_data"
    version = "0.1.0"

    def run(self, ctx: BlockContext, inputs: Dict[str, Any]) -> Dict[str, Any]:
        """Excelワークブックからデータを読み取り、行配列に整形して返す。

        入力形式:
          - workbook: { name: str, bytes: b"..." } または bytes
          - read_config:
              sheets: [ { name: str, header_row?: int, range?: str, skip_empty_rows?: bool } ]
              header_row: 1 (既定)
              skip_empty_rows: true (既定)
              date_as_iso: true (既定)
        出力:
          - data: { <sheet>: [ {header: value, ...}, ... ] }
          - summary: { sheets: int, rows: {<sheet>: int} }
        """

        wb_input = inputs.get("workbook")
        wb_name: str | None = None
        wb_bytes: bytes | bytearray | None = None
        wb_path: Path | None = None
        if isinstance(wb_input, dict):
            wb_name = wb_input.get("name") if isinstance(wb_input.get("name"), str) else None
            if isinstance(wb_input.get("bytes"), (bytes, bytearray)):
                wb_bytes = wb_input.get("bytes")
            elif isinstance(wb_input.get("path"), str):
                wb_path = Path(wb_input.get("path"))
        elif isinstance(wb_input, (bytes, bytearray)):
            wb_bytes = wb_input
        elif isinstance(wb_input, str):
            wb_path = Path(wb_input)

        if wb_path is None and not isinstance(wb_bytes, (bytes, bytearray)):
            return {"data": {}, "summary": {"sheets": 0, "rows": {}}}

        cfg = inputs.get("read_config") or {}
        # 読み取りモード: single | multi （single 既定。single時はトップレベル rows を併せて返す）
        mode_raw = cfg.get("mode") if isinstance(cfg, dict) else None
        if mode_raw is None:
            mode_raw = inputs.get("mode")
        mode = str(mode_raw or "single").strip().lower()
        default_header_row = int(cfg.get("header_row", 1) or 1)
        default_skip_empty = bool(cfg.get("skip_empty_rows", True))
        date_as_iso = bool(cfg.get("date_as_iso", True))

        # Optional: Recalculate via LibreOffice headless, then read computed results
        # Inputs:
        #   recalc: true | { enabled?: true, engine?: "libreoffice", soffice_path?: str, timeout_sec?: int }
        recalc_cfg = inputs.get("recalc")
        recalc_enabled = False
        recalc_engine: str | None = None  # 'libreoffice' | 'pycel'
        soffice_path: str | None = None
        timeout_sec: int = 120
        if isinstance(recalc_cfg, bool):
            recalc_enabled = recalc_cfg
            recalc_engine = "libreoffice" if recalc_enabled else None
        elif isinstance(recalc_cfg, str):
            low = recalc_cfg.lower()
            if low in {"pycel"}:
                recalc_enabled = True
                recalc_engine = "pycel"
            else:
                recalc_enabled = low in {"true", "1", "yes", "y", "on", "libreoffice"}
                recalc_engine = "libreoffice" if recalc_enabled else None
        elif isinstance(recalc_cfg, dict):
            recalc_enabled = bool(recalc_cfg.get("enabled", True))
            eng = recalc_cfg.get("engine")
            if isinstance(eng, str) and eng.lower() in {"pycel", "libreoffice"}:
                recalc_engine = eng.lower()
            else:
                recalc_engine = "libreoffice"
            soffice_path = recalc_cfg.get("soffice_path") if isinstance(recalc_cfg.get("soffice_path"), str) else None
            try:
                timeout_sec = int(recalc_cfg.get("timeout_sec", 120))
            except Exception:
                timeout_sec = 120
        else:
            recalc_engine = None

        recalc_pycel = bool(recalc_enabled and recalc_engine == "pycel")
        recalc_libreoffice = bool(recalc_enabled and (recalc_engine in (None, "libreoffice")))

        recalculated_bytes: bytes | None = None
        recalc_status: str = "skipped"
        if recalc_libreoffice:
            # Resolve soffice executable
            candidates: List[str] = []
            if soffice_path:
                candidates.append(soffice_path)
            env_path = os.environ.get("LIBREOFFICE_PATH")
            if env_path:
                candidates.append(env_path)
            # Common Windows install locations
            candidates.extend([
                "soffice",
                r"C:\\Program Files\\LibreOffice\\program\\soffice.exe",
                r"C:\\Program Files (x86)\\LibreOffice\\program\\soffice.exe",
            ])

            def _pick_soffice() -> str | None:
                for c in candidates:
                    try:
                        # If it's an absolute path, check existence; otherwise rely on PATH resolution
                        if os.path.isabs(c):
                            if os.path.exists(c):
                                return c
                        else:
                            # Try to resolve via shutil.which
                            w = shutil.which(c)
                            if w:
                                return w
                    except Exception:
                        continue
                return None

            soff = _pick_soffice()
            if soff is not None:
                try:
                    with tempfile.TemporaryDirectory() as tmpdir:
                        tmpd = Path(tmpdir)
                        # Always work on a temp copy to avoid touching original
                        in_path = tmpd / (wb_name or "input.xlsx")
                        if wb_path is not None and wb_path.exists():
                            try:
                                shutil.copyfile(str(wb_path), str(in_path))
                            except Exception:
                                in_path.write_bytes(wb_path.read_bytes())
                        else:
                            try:
                                in_path.write_bytes(bytes(wb_bytes or b""))
                            except Exception:
                                in_path.write_bytes(b"")
                        out_dir = tmpd
                        # Pass 1: convert to ODS (forces LO to open and save)
                        cmd1 = [soff, "--headless", "--convert-to", "ods", "--outdir", str(out_dir), str(in_path)]
                        proc1 = subprocess.run(cmd1, stdout=subprocess.PIPE, stderr=subprocess.PIPE, timeout=timeout_sec)
                        ods_path = out_dir / (in_path.stem + ".ods")
                        if not ods_path.exists():
                            recalc_status = f"convert_to_ods_missing_output (rc={proc1.returncode})"
                        else:
                            # Optionally rename to stable basename
                            recalc_ods = out_dir / "recalc.ods"
                            try:
                                if recalc_ods.exists():
                                    recalc_ods.unlink()
                                ods_path.rename(recalc_ods)
                            except Exception:
                                recalc_ods = ods_path
                            # Pass 2: convert ODS back to XLSX (re-save with recalculated values)
                            # Note: explicit filter name improves reliability on some installs
                            cmd2 = [
                                soff,
                                "--headless",
                                "--convert-to",
                                "xlsx:Calc MS Excel 2007 XML",
                                "--outdir",
                                str(out_dir),
                                str(recalc_ods),
                            ]
                            proc2 = subprocess.run(cmd2, stdout=subprocess.PIPE, stderr=subprocess.PIPE, timeout=timeout_sec)
                            out_path = out_dir / (recalc_ods.stem + ".xlsx")
                            if out_path.exists():
                                recalculated_bytes = out_path.read_bytes()
                                recalc_status = "ok_2pass"
                            else:
                                recalc_status = f"convert_to_xlsx_missing_output (rc={proc2.returncode})"
                except subprocess.TimeoutExpired:
                    recalc_status = "timeout"
                except Exception as _:
                    recalc_status = "failed"
            else:
                recalc_status = "soffice_not_found"

        # If LibreOffice recalc was requested but failed, raise an error (do not silently fallback)
        if recalc_libreoffice and recalculated_bytes is None:
            code = ErrorCode.EXTERNAL_TIMEOUT if recalc_status == "timeout" else ErrorCode.EXTERNAL_API_ERROR
            raise BlockException(
                BlockError(
                    code=code,
                    message="LibreOffice headless recalc failed",
                    details={
                        "status": recalc_status,
                        "soffice_path": soff if 'soff' in locals() else None,
                    },
                    hint=(
                        "LibreOfficeが見つからない、または変換に失敗しました。LibreOfficeのインストールとPATH設定を確認するか、"
                        "recalc: { soffice_path: 'C:/Program Files/LibreOffice/program/soffice.exe' } を指定してください。"
                    ),
                    recoverable=False,
                )
            )

        # Decide source bytes to load
        # Prepare workbook for reading
        if recalc_pycel:
            # For pycel, we need a file path. Persist to a temp file when necessary.
            tmp_file_path: Path | None = None
            pycel_input_path: Path
            if wb_path is not None and wb_path.exists():
                pycel_input_path = wb_path
            else:
                tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
                try:
                    tmp.write(bytes(wb_bytes or b""))
                finally:
                    tmp.close()
                tmp_file_path = Path(tmp.name)
                pycel_input_path = tmp_file_path
            # Load openpyxl workbook with formulas (data_only=False)
            wb = load_workbook(str(pycel_input_path), data_only=False)
            # Build pycel compiler
            try:
                from pycel.excelcompiler import ExcelCompiler  # type: ignore
            except Exception as e:
                # Cleanup temp file before raising
                if tmp_file_path is not None:
                    try:
                        os.unlink(str(tmp_file_path))
                    except Exception:
                        pass
                raise BlockException(
                    BlockError(
                        code=ErrorCode.DEPENDENCY_NOT_FOUND,
                        message="pycel is not installed",
                        details={"import_error": str(e)},
                        hint="pip install pycel を実行してください。",
                        recoverable=False,
                    )
                )
            try:
                xlc = ExcelCompiler(filename=str(pycel_input_path))  # type: ignore
                recalc_status = "pycel_ok"
            except Exception as e:
                if tmp_file_path is not None:
                    try:
                        os.unlink(str(tmp_file_path))
                    except Exception:
                        pass
                raise BlockException(
                    BlockError(
                        code=ErrorCode.EXTERNAL_API_ERROR,
                        message="pycel ExcelCompiler failed to load",
                        details={"path": str(pycel_input_path), "error": str(e)},
                        hint="対象のExcelがpycel非対応の関数/参照を含む可能性があります。",
                        recoverable=False,
                    )
                )
        else:
            src_bytes: bytes | bytearray | None = recalculated_bytes if recalculated_bytes is not None else wb_bytes
            if src_bytes is not None:
                wb = load_workbook(BytesIO(src_bytes), data_only=True)
            elif wb_path is not None and wb_path.exists():
                wb = load_workbook(str(wb_path), data_only=True)
            else:
                return {"data": {}, "summary": {"sheets": 0, "rows": {}}}

        # シート選択
        sheets_cfg = cfg.get("sheets")
        sheet_names: List[str]
        per_sheet_cfg: Dict[str, Dict[str, Any]] = {}
        if isinstance(sheets_cfg, list) and sheets_cfg:
            names: List[str] = []
            for s in sheets_cfg:
                if isinstance(s, dict) and isinstance(s.get("name"), str):
                    names.append(s["name"])
                    per_sheet_cfg[s["name"]] = s
            sheet_names = [n for n in names if n in wb.sheetnames]
        else:
            sheet_names = list(wb.sheetnames)

        out: Dict[str, List[Dict[str, Any]]] = {}
        rows_count: Dict[str, int] = {}

        for name in sheet_names:
            ws = wb[name]
            scfg = per_sheet_cfg.get(name, {})
            header_row = int(scfg.get("header_row", default_header_row) or default_header_row)
            skip_empty = bool(scfg.get("skip_empty_rows", default_skip_empty))
            rng = scfg.get("range")

            # 範囲決定
            use_values_only = True
            if rng and isinstance(rng, str):
                data_range = ws[rng]
                use_values_only = False  # ws[rng] は Cell オブジェクトを返す
            else:
                if recalc_pycel:
                    data_range = ws.iter_rows()
                    use_values_only = False
                else:
                    data_range = ws.iter_rows(values_only=True)
                    use_values_only = True

            headers: List[str] = []
            records: List[Dict[str, Any]] = []

            # 値イテレーション（iter_rowsの場合は全行を順に処理）
            current_row = 0
            for row in data_range:
                current_row += 1
                # openpyxlの iter_rows(values_only=True) はタプル
                cells = list(row) if isinstance(row, (list, tuple)) else [row]
                if not use_values_only:
                    # Cell -> 値 へ変換（pycel時は数式セルを評価）
                    if recalc_pycel:
                        evaled: List[Any] = []
                        first_error: Exception | None = None
                        failing_ref: str | None = None
                        for c in cells:
                            val = getattr(c, "value", None)
                            if isinstance(val, str) and val.startswith("="):
                                ref = f"{name}!{getattr(c, 'coordinate', '')}"
                                try:
                                    res = xlc.evaluate(ref)  # type: ignore[name-defined]
                                except Exception as ee:
                                    first_error = ee
                                    failing_ref = ref
                                    break
                                evaled.append(res)
                            else:
                                evaled.append(val)
                        if first_error is not None:
                            raise BlockException(
                                BlockError(
                                    code=ErrorCode.EXTERNAL_API_ERROR,
                                    message="pycel evaluation failed",
                                    details={"cell": failing_ref, "error": str(first_error)},
                                    hint="pycelが未対応の関数/参照、または循環参照が存在する可能性があります。",
                                    recoverable=False,
                                )
                            )
                        cells = evaled
                    else:
                        cells = [getattr(c, "value", c) for c in cells]
                if current_row == header_row:
                    headers = [str(c) if c is not None else f"col{idx+1}" for idx, c in enumerate(cells)]
                    continue
                if current_row < header_row:
                    continue
                if not headers:
                    # ヘッダが未確定の場合はスキップ
                    continue
                # 空行判定
                if skip_empty and all(c is None or (isinstance(c, str) and not c.strip()) for c in cells):
                    continue
                rec: Dict[str, Any] = {}
                for idx, h in enumerate(headers):
                    val = cells[idx] if idx < len(cells) else None
                    # 日付→ISO
                    if date_as_iso and isinstance(val, datetime):
                        val = val.date().isoformat()
                    rec[h] = val
                records.append(rec)

            out[name] = records
            rows_count[name] = len(records)

        summary = {"sheets": len(out), "rows": rows_count, "recalc": {"enabled": bool(recalc_enabled), "status": recalc_status}, "mode": mode}
        # ログ: 読み取り概要
        try:
            export_log({"sheets": list(out.keys()), "rows": rows_count, "recalc": recalc_status}, ctx=ctx, tag="excel.read_data")
        except Exception:
            pass

        # singleモード時はトップレベル rows を追加（厳密には Plan で1枚のみを想定）
        if mode == "single":
            selected_rows: List[Dict[str, Any]] = []
            try:
                sheets_cfg = cfg.get("sheets")
                if isinstance(sheets_cfg, list) and len(sheets_cfg) == 1 and isinstance(sheets_cfg[0], dict):
                    sname = sheets_cfg[0].get("name")
                    if isinstance(sname, str) and sname in out:
                        selected_rows = out.get(sname, [])
                # フォールバック: 1枚だけ読み取れた場合
                if not selected_rows and len(out.keys()) == 1:
                    only = next(iter(out.keys()))
                    selected_rows = out.get(only, [])
            except Exception:
                selected_rows = []
            return {"data": out, "rows": selected_rows, "summary": summary}

        return {"data": out, "summary": summary}



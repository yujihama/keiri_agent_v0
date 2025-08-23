from __future__ import annotations

from typing import Any, Dict, List
from io import BytesIO
import base64

from openpyxl import load_workbook, Workbook
from openpyxl.utils.cell import coordinate_to_tuple
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string
from openpyxl.utils.cell import range_boundaries
from pathlib import Path

from core.blocks.base import BlockContext, ProcessingBlock


class ExcelUpdateWorkbookBlock(ProcessingBlock):
    id = "excel.update_workbook"
    version = "0.1.0"

    def run(self, ctx: BlockContext, inputs: Dict[str, Any]) -> Dict[str, Any]:
        """命令配列に基づいて既存Excelワークブックを更新する。"""

        wb_input = inputs.get("workbook")
        wb_name = wb_input.get("name") if isinstance(wb_input, dict) else None
        wb_bytes_or_path = wb_input.get("bytes") if isinstance(wb_input, dict) else wb_input

        wb: Workbook
        if isinstance(wb_bytes_or_path, (bytes, bytearray)):
            wb = load_workbook(BytesIO(wb_bytes_or_path))
        elif isinstance(wb_bytes_or_path, str):
            # ファイルパス指定を許容
            p = Path(wb_bytes_or_path)
            if p.exists() and p.is_file():
                wb = load_workbook(p)
            else:
                wb = Workbook()
        else:
            # 空ブックで開始
            wb = Workbook()

        ops = inputs.get("operations") or []
        if isinstance(ops, dict):
            ops = [ops]

        def _get_ws(name: str) -> Worksheet:
            if name in wb.sheetnames:
                return wb[name]
            return wb.create_sheet(name)

        def _append_table(ws: Worksheet, start_addr: str, data: Any) -> int:
            # data: list[dict] | dict[str, list] | DataFrame
            def _as_records(values_obj: Any) -> List[Dict[str, Any]]:
                try:
                    import pandas as _pd  # type: ignore
                except Exception:
                    _pd = None
                if _pd is not None and isinstance(values_obj, _pd.DataFrame):
                    return values_obj.to_dict(orient="records")
                if isinstance(values_obj, list) and all(isinstance(x, dict) for x in values_obj):
                    return values_obj
                if isinstance(values_obj, dict):
                    keys = list(values_obj.keys())
                    try:
                        max_len = max(len(values_obj[k]) for k in keys)
                    except Exception:
                        return []
                    out: List[Dict[str, Any]] = []
                    for i in range(max_len):
                        rec: Dict[str, Any] = {}
                        for k in keys:
                            seq = values_obj.get(k)
                            rec[k] = seq[i] if isinstance(seq, list) and i < len(seq) else None
                        out.append(rec)
                    return out
                return []

            try:
                r0, c0 = coordinate_to_tuple(start_addr)
            except Exception:
                r0, c0 = 1, 1
            records = _as_records(data)
            # ヘッダ書き込み
            headers = list(records[0].keys()) if records else []
            for j, h in enumerate(headers, start=c0):
                ws.cell(row=r0, column=j, value=h)
            # データ
            r = r0 + 1
            for rec in records:
                for j, h in enumerate(headers, start=c0):
                    ws.cell(row=r, column=j, value=rec.get(h))
                r += 1
            return max(0, len(records))

        def _copy_conditional_formatting(src_ws: Worksheet, dst_ws: Worksheet) -> None:
            """src_ws の条件付き書式を dst_ws に複製する。
            openpyxl のバージョン差異に配慮して rules/_cf_rules の双方を試行する。
            重複は簡易シグネチャで抑止する。
            """
            try:
                src_cf = getattr(src_ws, "conditional_formatting", None)
                dst_cf = getattr(dst_ws, "conditional_formatting", None)
                if src_cf is None or dst_cf is None:
                    return

                # 既存の宛先ルールの簡易シグネチャを構築
                existing: set = set()
                try:
                    dest_rules_map = getattr(dst_cf, "rules", None)
                    if isinstance(dest_rules_map, dict):
                        for rdef, rules in dest_rules_map.items():
                            for rule in rules:
                                existing.add((str(rdef), getattr(rule, "type", None), tuple(getattr(rule, "formula", []) or [])))
                    else:
                        dest_cf_rules = getattr(dst_cf, "_cf_rules", None)
                        if isinstance(dest_cf_rules, dict):
                            for rdef, rules in dest_cf_rules.items():
                                for rule in rules:
                                    existing.add((str(rdef), getattr(rule, "type", None), tuple(getattr(rule, "formula", []) or [])))
                except Exception:
                    existing = set()

                # ルールの列挙（rules -> _cf_rules の順でフォールバック）
                def _iter_src_cf():
                    src_rules_map = getattr(src_cf, "rules", None)
                    if isinstance(src_rules_map, dict):
                        for rdef, rules in src_rules_map.items():
                            for rule in rules:
                                yield (str(rdef), rule)
                        return
                    src_cf_rules = getattr(src_cf, "_cf_rules", None)
                    if isinstance(src_cf_rules, dict):
                        for rdef, rules in src_cf_rules.items():
                            for rule in rules:
                                yield (str(rdef), rule)

                from copy import copy as _copy

                for rdef, rule in list(_iter_src_cf()):
                    sig = (str(rdef), getattr(rule, "type", None), tuple(getattr(rule, "formula", []) or []))
                    if sig in existing:
                        continue
                    try:
                        # まずはディープコピーで追加
                        dst_cf.add(str(rdef), _copy(rule))
                    except Exception:
                        try:
                            # コピー不可な場合は参照のまま（openpyxl が必要なら内部でコピーする）
                            dst_cf.add(str(rdef), rule)
                        except Exception:
                            # 最後に諦める
                            pass
            except Exception:
                # 何らかの理由で条件付き書式の複製に失敗しても全体処理は継続
                return

        total_updates = 0
        total_formats = 0
        total_rows_updated = 0
        for op in ops:
            if not isinstance(op, dict):
                continue
            optype = (op.get("type") or "").strip()
            if optype == "add_sheet":
                name = op.get("sheet_name") or "Sheet1"
                _get_ws(name)
            elif optype == "copy_sheet":
                src = op.get("sheet_name")
                dst = op.get("target")
                if isinstance(src, str) and isinstance(dst, str) and src in wb.sheetnames:
                    src_ws = wb[src]
                    new_ws = wb.copy_worksheet(src_ws)
                    new_ws.title = dst
                    # 条件付き書式の複製（openpyxl の copy_worksheet では失われる場合があるため補完）
                    try:
                        _copy_conditional_formatting(src_ws, new_ws)
                    except Exception:
                        pass
                    # placement control
                    idx = op.get("index")
                    pos = (op.get("position") or "").strip().lower()
                    before = op.get("before")
                    after = op.get("after")
                    try:
                        if isinstance(idx, int):
                            wb._sheets.remove(new_ws)
                            wb._sheets.insert(max(0, idx), new_ws)
                        elif pos == "first":
                            wb._sheets.remove(new_ws)
                            wb._sheets.insert(0, new_ws)
                        elif pos == "last":
                            pass  # default is last
                        elif isinstance(before, str) and before in wb.sheetnames:
                            wb._sheets.remove(new_ws)
                            tgt = wb[before]
                            i = wb._sheets.index(tgt)
                            wb._sheets.insert(i, new_ws)
                        elif isinstance(after, str) and after in wb.sheetnames:
                            wb._sheets.remove(new_ws)
                            tgt = wb[after]
                            i = wb._sheets.index(tgt)
                            wb._sheets.insert(i + 1, new_ws)
                    except Exception:
                        pass
            elif optype == "move_sheet":
                name = op.get("sheet_name")
                if isinstance(name, str) and name in wb.sheetnames:
                    ws = wb[name]
                    idx = op.get("index")
                    pos = (op.get("position") or "").strip().lower()
                    before = op.get("before")
                    after = op.get("after")
                    try:
                        wb._sheets.remove(ws)
                        if isinstance(idx, int):
                            wb._sheets.insert(max(0, idx), ws)
                        elif pos == "first":
                            wb._sheets.insert(0, ws)
                        elif pos == "last":
                            wb._sheets.append(ws)
                        elif isinstance(before, str) and before in wb.sheetnames:
                            tgt = wb[before]
                            i = wb._sheets.index(tgt)
                            wb._sheets.insert(i, ws)
                        elif isinstance(after, str) and after in wb.sheetnames:
                            tgt = wb[after]
                            i = wb._sheets.index(tgt)
                            wb._sheets.insert(i + 1, ws)
                        else:
                            wb._sheets.append(ws)
                    except Exception:
                        pass
            elif optype == "update_cells":
                name = op.get("sheet_name") or "Results"
                ws = _get_ws(name)
                cells = op.get("cells") or {}
                # 互換: target + data 指定を cells に変換
                if not cells and op.get("target") is not None:
                    cells = {str(op.get("target")): op.get("data")}
                for addr, val in list(cells.items()):
                    try:
                        r, c = coordinate_to_tuple(str(addr))
                        ws.cell(row=r, column=c, value=val)
                        total_updates += 1
                    except Exception:
                        continue
            elif optype == "append_table":
                name = op.get("sheet_name") or "Results"
                ws = _get_ws(name)
                target = op.get("target") or "A1"
                total_updates += _append_table(ws, str(target), op.get("data"))
            elif optype == "append_rows_bottom":
                # ワークシートの最終データ行の直下に、列レター指定で値を追加
                name = op.get("sheet_name") or "Results"
                ws = _get_ws(name)
                rows = op.get("rows") or []
                col_map = op.get("columns") or {}
                # 末尾行の検出: 指定列群（無ければ全列）で最終非空セルの行
                def _last_used_row() -> int:
                    scan_cols = []
                    if isinstance(col_map, dict):
                        for v in col_map.values():
                            if isinstance(v, str) and v:
                                scan_cols.append(v)
                    if not scan_cols:
                        # all columns as fallback
                        return ws.max_row or 1
                    from openpyxl.utils import column_index_from_string as _colidx
                    last = 0
                    for letter in scan_cols:
                        try:
                            cidx = _colidx(letter)
                        except Exception:
                            continue
                        for r in range(ws.max_row, 0, -1):
                            if ws.cell(row=r, column=cidx).value not in (None, ""):
                                last = max(last, r)
                                break
                    return max(1, last)
                start_row = _last_used_row() + 1
                from openpyxl.utils import column_index_from_string as _colidx
                for offset, item in enumerate(rows if isinstance(rows, list) else []):
                    r = start_row + offset
                    if isinstance(item, dict):
                        for field, letter in col_map.items():
                            try:
                                cidx = _colidx(str(letter))
                            except Exception:
                                continue
                            val = item.get(field)
                            try:
                                ws.cell(row=r, column=cidx, value=val)
                                total_updates += 1
                            except Exception:
                                continue
                    elif isinstance(item, (list, tuple)):
                        # listの場合、columnsの順に投入（field=0,1... を想定）
                        for i, letter in enumerate(col_map.values()):
                            try:
                                cidx = _colidx(str(letter))
                            except Exception:
                                continue
                            val = item[i] if i < len(item) else None
                            try:
                                ws.cell(row=r, column=cidx, value=val)
                                total_updates += 1
                            except Exception:
                                continue
            elif optype == "insert_rows":
                name = op.get("sheet_name") or "Results"
                ws = _get_ws(name)
                start_row = int(op.get("start_row") or 2)
                count = int(op.get("count") or 1)
                try:
                    ws.insert_rows(start_row, amount=count)
                except Exception:
                    pass
            elif optype == "update_formula":
                name = op.get("sheet_name") or "Results"
                ws = _get_ws(name)
                cells = op.get("cells") or {}
                for addr, formula in list(cells.items()):
                    try:
                        r, c = coordinate_to_tuple(str(addr))
                        ws.cell(row=r, column=c, value=str(formula))
                        total_updates += 1
                    except Exception:
                        continue
            elif optype == "update_cells_if":
                # 条件が真の時のみ update_cells を実行
                cond = bool(op.get("condition", True))
                if not cond:
                    continue
                name = op.get("sheet_name") or "Results"
                ws = _get_ws(name)
                cells = op.get("cells") or {}
                if not cells and op.get("target") is not None:
                    cells = {str(op.get("target")): op.get("data")}
                for addr, val in list(cells.items()):
                    try:
                        r, c = coordinate_to_tuple(str(addr))
                        ws.cell(row=r, column=c, value=val)
                        total_updates += 1
                    except Exception:
                        continue
            elif optype == "update_formula_range":
                # 範囲へ同一/テンプレートの数式を一括適用
                cond = bool(op.get("condition", True))
                if not cond:
                    continue
                name = op.get("sheet_name") or "Results"
                ws = _get_ws(name)
                rng = str(op.get("range") or "").strip()
                template = str(op.get("template") or "")
                if not rng or not template:
                    continue
                def _iter_cells(ws: Worksheet, rdef: str):
                    # サポート: "A1:C10", "K:K", "J3"
                    import re as _re
                    try:
                        if _re.fullmatch(r"[A-Za-z]+:[A-Za-z]+", rdef):
                            from openpyxl.utils import column_index_from_string as _colidx
                            left, right = rdef.split(":", 1)
                            c1, c2 = _colidx(left), _colidx(right)
                            if c1 > c2:
                                c1, c2 = c2, c1
                            for col in range(c1, c2 + 1):
                                for row in range(1, ws.max_row + 1):
                                    yield (row, col)
                            return
                        # 通常レンジ
                        min_col, min_row, max_col, max_row = range_boundaries(rdef)
                        for r in range(min_row, max_row + 1):
                            for c in range(min_col, max_col + 1):
                                yield (r, c)
                    except Exception:
                        # 単一セル
                        try:
                            rr, cc = coordinate_to_tuple(str(rdef))
                            yield (rr, cc)
                        except Exception:
                            return
                for (rr, cc) in _iter_cells(ws, rng):
                    try:
                        ws.cell(row=rr, column=cc, value=str(template))
                        total_updates += 1
                    except Exception:
                        continue
            elif optype == "replace_in_formulas":
                # 既存の数式文字列に対して部分置換を行う
                cond = bool(op.get("condition", True))
                if not cond:
                    continue
                name = op.get("sheet_name") or "Results"
                ws = _get_ws(name)
                rng = str(op.get("range") or "").strip()
                search = op.get("search")
                replace = op.get("replace")
                use_regex = bool(op.get("regex", False))
                match_case = bool(op.get("match_case", True))
                if not rng or search is None or replace is None:
                    continue
                import re
                flags = 0 if match_case else re.IGNORECASE
                if use_regex:
                    try:
                        pattern = re.compile(str(search), flags=flags)
                    except Exception:
                        pattern = None
                else:
                    pattern = None
                def _iter_cells(ws: Worksheet, rdef: str):
                    import re as _re
                    try:
                        if _re.fullmatch(r"[A-Za-z]+:[A-Za-z]+", rdef):
                            from openpyxl.utils import column_index_from_string as _colidx
                            left, right = rdef.split(":", 1)
                            c1, c2 = _colidx(left), _colidx(right)
                            if c1 > c2:
                                c1, c2 = c2, c1
                            for col in range(c1, c2 + 1):
                                for row in range(1, ws.max_row + 1):
                                    yield (row, col)
                            return
                        min_col, min_row, max_col, max_row = range_boundaries(rdef)
                        for r in range(min_row, max_row + 1):
                            for c in range(min_col, max_col + 1):
                                yield (r, c)
                    except Exception:
                        try:
                            rr, cc = coordinate_to_tuple(str(rdef))
                            yield (rr, cc)
                        except Exception:
                            return
                for (rr, cc) in _iter_cells(ws, rng):
                    try:
                        cell = ws.cell(row=rr, column=cc)
                        val = cell.value
                        if val is None:
                            continue
                        sval = str(val)
                        new_val = sval
                        if pattern is not None:
                            new_val = pattern.sub(str(replace), sval)
                        else:
                            if match_case:
                                new_val = sval.replace(str(search), str(replace))
                            else:
                                # 大文字小文字無視での置換
                                new_val = re.sub(re.escape(str(search)), str(replace), sval, flags=re.IGNORECASE)
                        if new_val != sval:
                            cell.value = new_val
                            total_updates += 1
                    except Exception:
                        continue
            elif optype == "clear_cells":
                # セルの値/数式/キャッシュをクリア
                name = op.get("sheet_name") or "Results"
                ws = _get_ws(name)
                targets = op.get("targets") or []
                addrs = targets if isinstance(targets, list) else [targets]
                for addr in addrs:
                    try:
                        r, c = coordinate_to_tuple(str(addr))
                        cell = ws.cell(row=r, column=c)
                        # 値と数式の両方をクリア
                        try:
                            cell.value = None
                        except Exception:
                            pass
                        # openpyxl内部フィールドのキャッシュも念のため除去
                        for attr in ("_value", "_formula", "_shared_formula"):
                            if hasattr(cell, attr):
                                try:
                                    setattr(cell, attr, None)
                                except Exception:
                                    pass
                        total_updates += 1
                    except Exception:
                        continue
            elif optype == "clear_cells_if":
                # 条件が真の場合のみ clear_cells を実行
                cond = op.get("condition")
                do_clear = bool(cond)  # すでに解決済みの値が入る想定
                if do_clear:
                    name = op.get("sheet_name") or "Results"
                    ws = _get_ws(name)
                    targets = op.get("targets") or []
                    addrs = targets if isinstance(targets, list) else [targets]
                    for addr in addrs:
                        try:
                            r, c = coordinate_to_tuple(str(addr))
                            cell = ws.cell(row=r, column=c)
                            cell.value = None
                            for attr in ("_value", "_formula", "_shared_formula"):
                                if hasattr(cell, attr):
                                    try:
                                        setattr(cell, attr, None)
                                    except Exception:
                                        pass
                            total_updates += 1
                        except Exception:
                            continue
            elif optype == "format_cells":
                # セル/範囲の塗りつぶし等のフォーマット
                name = op.get("sheet_name") or "Results"
                ws = _get_ws(name)
                ranges = op.get("ranges") or []
                fill_cfg = op.get("fill") or {}
                color = (fill_cfg.get("color") or "D9D9D9").lstrip("#").upper()
                if len(color) == 6:
                    color = "FF" + color
                fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                for rdef in ranges if isinstance(ranges, list) else [ranges]:
                    if not isinstance(rdef, str) or not rdef:
                        continue
                    try:
                        min_col, min_row, max_col, max_row = range_boundaries(rdef)
                    except Exception:
                        # 単一セルの場合のフォールバック
                        try:
                            rr, cc = coordinate_to_tuple(str(rdef))
                            min_row = max_row = rr
                            min_col = max_col = cc
                        except Exception:
                            continue
                    for r in range(min_row, max_row + 1):
                        for c in range(min_col, max_col + 1):
                            try:
                                ws.cell(row=r, column=c).fill = fill
                                total_formats += 1
                            except Exception:
                                continue
            elif optype == "update_rows_by_match":
                # ヘッダ行を検出し、キー列の値で行を特定して更新・着色
                name = op.get("sheet_name") or "Results"
                ws = _get_ws(name)
                header_row_opt = op.get("header_row")
                key_header = op.get("key") or op.get("match_header")
                key_column_letter = op.get("key_column") or op.get("match_column")  # 列レターが優先
                items = op.get("items") or []
                update_fields = op.get("update_fields") or {}
                update_columns = op.get("update_columns") or {}
                # 例: "G:X"
                fill_range_cols = op.get("fill_range_columns")
                fill_color = (op.get("fill_color") or "D9D9D9").lstrip("#").upper()
                if len(fill_color) == 6:
                    fill_color = "FF" + fill_color
                fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

                # ヘッダ行の自動検出
                def detect_header_row() -> int:
                    if isinstance(header_row_opt, int) and header_row_opt > 0:
                        return header_row_opt
                    if key_header:
                        # 先頭50行程度からキー見出しを探索
                        for r in range(1, min(51, ws.max_row + 1)):
                            names = []
                            for c in range(1, ws.max_column + 1):
                                v = ws.cell(row=r, column=c).value
                                if v is None:
                                    names.append("")
                                else:
                                    names.append(str(v).strip())
                            if any(n for n in names) and key_header in names:
                                return r
                    return 1

                header_row = detect_header_row()
                # 見出し→列番号（または列レター）
                header_map: Dict[str, int] = {}
                if not key_column_letter:
                    for c in range(1, ws.max_column + 1):
                        v = ws.cell(row=header_row, column=c).value
                        if v is not None:
                            header_map[str(v).strip()] = c
                    key_col = header_map.get(str(key_header).strip())
                else:
                    try:
                        key_col = column_index_from_string(str(key_column_letter))
                    except Exception:
                        key_col = None
                if not key_col:
                    # 見つからなければスキップ
                    pass
                else:
                    # 列範囲の解析
                    fill_start_col = fill_end_col = None
                    if isinstance(fill_range_cols, str) and ":" in fill_range_cols:
                        left, right = [s.strip() for s in fill_range_cols.split(":", 1)]
                        try:
                            fill_start_col = column_index_from_string(left)
                            fill_end_col = column_index_from_string(right)
                            if fill_start_col > fill_end_col:
                                fill_start_col, fill_end_col = fill_end_col, fill_start_col
                        except Exception:
                            fill_start_col = fill_end_col = None

                    for item in items if isinstance(items, list) else []:
                        try:
                            key_val = item.get(key_header)
                        except Exception:
                            key_val = None
                        target_row = None
                        if key_val is not None:
                            for r in range(header_row + 1, ws.max_row + 1):
                                cv = ws.cell(row=r, column=key_col).value
                                if cv is None:
                                    continue
                                if str(cv).strip() == str(key_val).strip():
                                    target_row = r
                                    break
                        # 氏名など別キーでのフォールバックは提供しない（汎用化）
                        if target_row is None:
                            continue

                        # 値の更新（ヘッダ指定 or 列レター指定）
                        if isinstance(update_fields, dict) and update_fields:
                            for src_field, dst_header in update_fields.items():
                                try:
                                    val = item.get(src_field)
                                except Exception:
                                    val = None
                                cidx = header_map.get(dst_header)
                                if cidx:
                                    try:
                                        ws.cell(row=target_row, column=cidx, value=val)
                                        total_rows_updated += 1
                                    except Exception:
                                        continue
                        if isinstance(update_columns, dict) and update_columns:
                            for src_field, col_letter in update_columns.items():
                                try:
                                    val = item.get(src_field)
                                except Exception:
                                    val = None
                                try:
                                    cidx = column_index_from_string(str(col_letter))
                                except Exception:
                                    continue
                                try:
                                    ws.cell(row=target_row, column=cidx, value=val)
                                    total_rows_updated += 1
                                except Exception:
                                    continue
                        # 行の着色
                        if fill_start_col is not None and fill_end_col is not None:
                            for c in range(fill_start_col, fill_end_col + 1):
                                try:
                                    ws.cell(row=target_row, column=c).fill = fill
                                    total_formats += 1
                                except Exception:
                                    continue
            else:
                # 未知タイプは無視
                continue

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        out_bytes = buf.getvalue()
        return {
            "workbook_updated": {"name": wb_name, "bytes": out_bytes},
            "workbook_b64": base64.b64encode(out_bytes).decode("ascii"),
            "summary": {
                "operations": len(ops),
                "cells_updated": total_updates,
                "cells_formatted": total_formats,
                "rows_updated": total_rows_updated,
            },
        }



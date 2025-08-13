from __future__ import annotations

from typing import Any, Dict, List
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.utils.cell import coordinate_to_tuple
from openpyxl.utils import column_index_from_string
from openpyxl.styles import PatternFill
import base64

from core.blocks.base import BlockContext, ProcessingBlock


class ExcelWriteBlock(ProcessingBlock):
    """Unified Excel writer.

    Official planner-facing block for writing results to Excel with a minimal,
    unambiguous schema.

    Inputs (exactly one of the following should be provided):
      - cell_updates: { sheet?: str, cells: { "A1": any, ... } } | [ {...}, ... ]
      - column_updates: {
            sheet?: str,
            header_row?: int (default 1),
            start_row?: int (default 2),
            columns: [ { header: str, path: str }, ... ],
            values: list | dict[str, list] | pandas.DataFrame,
            write_header?: bool (default true),
            clear_existing?: bool (default false)
        } | [ {...}, ... ]
      - workbook: { name?: str, bytes?: bytes } | bytes

    Outputs:
      - write_summary: { rows_written: int, sheet: str | null, workbook_name: str | None }
      - workbook_updated: { name: str | None, bytes: bytes }
      - workbook_b64: str
    """

    id = "excel.write"
    version = "0.1.0"

    def run(self, ctx: BlockContext, inputs: Dict[str, Any]) -> Dict[str, Any]:
        wb_name = None
        wb_bytes = None
        wb_input = inputs.get("workbook")
        if isinstance(wb_input, dict):
            wb_name = wb_input.get("name")
            wb_bytes = wb_input.get("bytes")
        elif isinstance(wb_input, (bytes, bytearray)):
            wb_bytes = wb_input

        # Load workbook; create new if invalid/missing
        try:
            if isinstance(wb_bytes, (bytes, bytearray)):
                wb: Workbook = load_workbook(BytesIO(wb_bytes))
            else:
                raise Exception("no-bytes")
        except Exception:
            from openpyxl import Workbook as _WB

            wb = _WB()

        def _get_ws(sheet: str | None, default: str = "Results"):
            sh = sheet or default
            if sh in wb.sheetnames:
                return wb[sh]
            if len(wb.sheetnames) == 1 and wb.active and wb.active.max_row == 1 and wb.active.max_column == 1:
                wb.active.title = sh
                return wb[sh]
            return wb.create_sheet(sh)

        def _as_records(values_obj: Any) -> List[Dict[str, Any]]:
            try:
                import pandas as _pd  # type: ignore
            except Exception:  # pragma: no cover
                _pd = None
            if _pd is not None and isinstance(values_obj, _pd.DataFrame):
                return values_obj.to_dict(orient="records")
            if isinstance(values_obj, list) and all(isinstance(x, dict) for x in values_obj):
                return values_obj
            if isinstance(values_obj, dict):
                keys = list(values_obj.keys())
                try:
                    max_len = max(len(values_obj[k]) for k in keys)
                except Exception:
                    return []
                out: List[Dict[str, Any]] = []
                for i in range(max_len):
                    rec: Dict[str, Any] = {}
                    for k in keys:
                        seq = values_obj.get(k)
                        rec[k] = seq[i] if isinstance(seq, list) and i < len(seq) else None
                    out.append(rec)
                return out
            return []

        def _get_by_path(obj: Any, path: str) -> Any:
            try:
                cur = obj
                for raw_seg in str(path).split("."):
                    seg = str(raw_seg).strip()
                    if not seg:
                        return None
                    if isinstance(cur, dict):
                        found = None
                        for k in cur.keys():
                            if str(k).lower() == seg.lower():
                                found = k
                                break
                        if found is None:
                            return None
                        cur = cur.get(found)
                    else:
                        return None
                return cur
            except Exception:
                return None

        def _deep_find_value(obj: Any, key_name: str, max_depth: int = 3) -> Any:
            try:
                if max_depth < 0:
                    return None
                if isinstance(obj, dict):
                    for k, v in obj.items():
                        if str(k).lower() == str(key_name).lower():
                            return v
                    for v in obj.values():
                        found = _deep_find_value(v, key_name, max_depth - 1)
                        if found is not None:
                            return found
                elif isinstance(obj, list):
                    for it in obj:
                        found = _deep_find_value(it, key_name, max_depth - 1)
                        if found is not None:
                            return found
            except Exception:
                return None
            return None

        def _resolve_value(rec: Dict[str, Any], key_or_path: str) -> Any:
            val = _get_by_path(rec, str(key_or_path))
            if val is not None:
                return val
            if "." not in str(key_or_path):
                for wrapper in ("row_data", "results"):
                    if isinstance(rec, dict) and wrapper in rec:
                        val = _get_by_path(rec, f"{wrapper}.{key_or_path}")
                        if val is not None:
                            return val
            return _deep_find_value(rec, str(key_or_path))

        def _col_index(col: Any, default_idx: int) -> int:
            if isinstance(col, int):
                return max(1, col)
            if isinstance(col, str):
                try:
                    return column_index_from_string(col)
                except Exception:
                    return default_idx
            return default_idx

        rows_written = 0
        last_sheet: str | None = None

        # 1) cell_updates
        cell_updates = inputs.get("cell_updates")
        if cell_updates is not None:
            ops: List[Dict[str, Any]] = []
            if isinstance(cell_updates, list):
                ops = [x for x in cell_updates if isinstance(x, dict)]
            elif isinstance(cell_updates, dict):
                ops = [cell_updates]
            for op in ops:
                sh = op.get("sheet")
                ws = _get_ws(sh)
                last_sheet = ws.title
                cells = op.get("cells") if isinstance(op.get("cells"), dict) else {}
                # also allow top-level A1 keys
                for k, v in list(op.items()):
                    try:
                        if isinstance(k, str) and k and k[0].isalpha():
                            r, c = coordinate_to_tuple(k)
                            cells.setdefault(k, v)
                    except Exception:
                        continue
                for addr, val in cells.items():
                    try:
                        r, c = coordinate_to_tuple(str(addr))
                    except Exception:
                        continue
                    try:
                        ws.cell(row=r, column=c, value=val)
                    except Exception:
                        continue
            # cell updates do not increment rows_written

        # 2) column_updates
        column_updates = inputs.get("column_updates")
        if column_updates is not None:
            ops2: List[Dict[str, Any]] = []
            if isinstance(column_updates, list):
                ops2 = [x for x in column_updates if isinstance(x, dict)]
            elif isinstance(column_updates, dict):
                ops2 = [column_updates]
            for op in ops2:
                sh = op.get("sheet")
                ws = _get_ws(sh)
                last_sheet = ws.title
                header_row = int(op.get("header_row") or 1)
                start_row = int(op.get("start_row") or 2)
                append = bool(op.get("append", False))
                write_header = bool(op.get("write_header", True))
                clear_existing = bool(op.get("clear_existing", False))
                columns_def = op.get("columns") or []
                values_obj = op.get("values")

                # normalize columns_def to list[{header, path}]
                cols_list: List[Dict[str, str]] = []
                if isinstance(columns_def, dict):
                    for header, path in columns_def.items():
                        if isinstance(path, str):
                            cols_list.append({"header": str(header), "path": path})
                elif isinstance(columns_def, list):
                    for item in columns_def:
                        if isinstance(item, dict):
                            header = item.get("header") or item.get("title") or item.get("name")
                            path = item.get("path") or item.get("field") or item.get("key") or item.get("name")
                            if isinstance(path, str) and path:
                                cols_list.append({"header": str(header or path.split(".")[-1]), "path": path})
                        elif isinstance(item, str):
                            cols_list.append({"header": item.split(".")[-1], "path": item})

                records = _as_records(values_obj)

                if clear_existing and cols_list:
                    try:
                        for idx in range(1, len(cols_list) + 1):
                            for rr in range(start_row, (ws.max_row or start_row) + 1):
                                ws.cell(row=rr, column=idx, value=None)
                    except Exception:
                        pass

                if write_header and cols_list:
                    for idx, coldef in enumerate(cols_list, start=1):
                        ws.cell(row=header_row, column=idx, value=coldef.get("header"))

                # append 対応: 末尾行の直下に書き込む
                if append:
                    try:
                        # スキャン対象列数（1..N）
                        col_count = max(1, len(cols_list))
                        last = 0
                        for c in range(1, col_count + 1):
                            for rr in range(ws.max_row, 0, -1):
                                if ws.cell(row=rr, column=c).value not in (None, ""):
                                    last = max(last, rr)
                                    break
                        start_r_eff = max(start_row, last + 1)
                    except Exception:
                        start_r_eff = start_row
                else:
                    start_r_eff = start_row

                r = start_r_eff
                for rec in records:
                    if not isinstance(rec, dict):
                        continue
                    for idx, coldef in enumerate(cols_list, start=1):
                        path = coldef.get("path") or ""
                        val = _resolve_value(rec, str(path))
                        ws.cell(row=r, column=idx, value=val)
                    r += 1
                    rows_written += 1

        # 3) match_updates (match + update + optional row fill)
        match_updates = inputs.get("match_updates")
        if match_updates is not None:
            ops3: List[Dict[str, Any]] = []
            if isinstance(match_updates, list):
                ops3 = [x for x in match_updates if isinstance(x, dict)]
            elif isinstance(match_updates, dict):
                ops3 = [match_updates]

            def _parse_range_columns(rdef: str) -> tuple[int | None, int | None]:
                try:
                    if ":" in rdef:
                        left, right = [s.strip() for s in rdef.split(":", 1)]
                        c1 = column_index_from_string(left)
                        c2 = column_index_from_string(right)
                        if c1 > c2:
                            c1, c2 = c2, c1
                        return c1, c2
                except Exception:
                    return None, None
                return None, None

            for op in ops3:
                sh = op.get("sheet")
                ws = _get_ws(sh)
                last_sheet = ws.title

                key_column = op.get("key_column")
                key_field = op.get("key_field")
                start_row = int(op.get("start_row") or 2)
                items = op.get("items") or []
                update_columns = op.get("update_columns") or {}
                fill_range = op.get("fill_range_columns")
                fill_color = (op.get("fill_color") or "D9D9D9").lstrip("#").upper()
                if len(fill_color) == 6:
                    fill_color = "FF" + fill_color

                try:
                    key_cidx = column_index_from_string(str(key_column))
                except Exception:
                    key_cidx = None

                if not key_cidx or not isinstance(key_field, str) or not key_field:
                    continue

                fill_c1: int | None = None
                fill_c2: int | None = None
                if isinstance(fill_range, str) and fill_range:
                    c1, c2 = _parse_range_columns(fill_range)
                    fill_c1, fill_c2 = c1, c2
                fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

                for it in items if isinstance(items, list) else []:
                    if not isinstance(it, dict):
                        continue
                    key_val = it.get(key_field)
                    if key_val is None:
                        continue
                    target_row = None
                    try:
                        s_key = str(key_val).strip()
                    except Exception:
                        s_key = None
                    if not s_key:
                        continue

                    for rr in range(start_row, (ws.max_row or start_row) + 1):
                        cv = ws.cell(row=rr, column=key_cidx).value
                        if cv is None:
                            continue
                        try:
                            if str(cv).strip() == s_key:
                                target_row = rr
                                break
                        except Exception:
                            continue

                    if target_row is None:
                        continue

                    # update columns
                    for src_field, col_letter in update_columns.items():
                        try:
                            cidx2 = column_index_from_string(str(col_letter))
                        except Exception:
                            continue
                        try:
                            ws.cell(row=target_row, column=cidx2, value=it.get(src_field))
                        except Exception:
                            continue

                    # optional row fill
                    if fill_c1 is not None and fill_c2 is not None:
                        for cc in range(fill_c1, fill_c2 + 1):
                            try:
                                ws.cell(row=target_row, column=cc).fill = fill
                            except Exception:
                                continue

        # Save
        out_buf = BytesIO()
        wb.save(out_buf)
        out_bytes = out_buf.getvalue()

        return {
            "write_summary": {"rows_written": rows_written, "sheet": last_sheet, "workbook_name": wb_name},
            "workbook_updated": {"name": wb_name, "bytes": out_bytes},
            "workbook_b64": base64.b64encode(out_bytes).decode("ascii"),
        }



from __future__ import annotations

from typing import Any, Dict, List
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.cell import coordinate_to_tuple
import base64

from core.blocks.base import BlockContext, ProcessingBlock


class ExcelWriteResultsBlock(ProcessingBlock):
    id = "excel.write_results"
    version = "0.1.0"

    def run(self, ctx: BlockContext, inputs: Dict[str, Any]) -> Dict[str, Any]:
        """Write match results into a provided workbook bytes.

        Inputs
        ------
        - workbook: { name: str, bytes: b"..." } or raw bytes
        - data: { matched: bool, items: [{file, count, sum}] }  (後方互換: 旧スタイル)
        - output_config: { sheet: str, start_row: int, columns: ["A","B","C", ...] } (後方互換: 旧スタイル)

        新スタイル（推奨）
        ------------------
        - cell_updates: 以下のいずれか
            1) { "sheet": str (任意), "cells": { "A1": any, "B2": any, ... } }
            2) [ { "sheet": str (任意), "cells": { "A1": any, ... } }, ... ]
            3) { "A1": any, "B2": any, ... }  # 既定シートに適用
        - column_updates: 以下のいずれか
            1) { "sheet": str (任意), "start_row": int (任意, 2既定), "column_map": {key: "A"|"B"|...}, "values": list|dict|DataFrame }
            2) [ 同上オブジェクト, ... ]
          values の受理形式:
            - list[dict]: 各行辞書（key は column_map のキー）
            - dict[str, list]: 列ごとの配列を行に整形
            - pandas.DataFrame: to_dict("records") 相当に変換

        Outputs
        -------
        - write_summary: { rows_written: int, sheet: str }
        """

        wb_bytes = None
        wb_name = None
        wb_input = inputs.get("workbook")
        if isinstance(wb_input, dict):
            wb_name = wb_input.get("name")
            wb_bytes = wb_input.get("bytes")
        elif isinstance(wb_input, (bytes, bytearray)):
            wb_bytes = wb_input

        if not isinstance(wb_bytes, (bytes, bytearray)):
            wb_bytes = None

        data = inputs.get("data") or {}
        # Normalize data: accept dict or JSON string; otherwise fallback to empty
        if isinstance(data, str):
            try:
                import json as _json
                data = _json.loads(data)
            except Exception:
                data = {}
        out_cfg = inputs.get("output_config") or {}
        # Normalize output_config as dict; if unresolved string, fallback to defaults
        if isinstance(out_cfg, str):
            out_cfg = {}
        sheet_name = out_cfg.get("sheet", "Results")

        def _coerce_int(val: Any, default_val: int) -> int:
            try:
                if val is None:
                    return default_val
                if isinstance(val, bool):
                    return default_val
                if isinstance(val, (int, float)):
                    return int(val)
                if isinstance(val, str):
                    s = val.strip()
                    return int(s) if s else default_val
            except Exception:
                return default_val
            return default_val

        start_row = _coerce_int(out_cfg.get("start_row"), 2)
        columns: List[str] = list(out_cfg.get("columns", ["A", "B", "C"]))
        # 列マッピング: キー名から列インデックスを解決可能に（列名指定優先、足りない場合は自動）
        # header_map = out_cfg.get("header_map") or {"file": "File", "count": "Count", "sum": "Sum"}

        cell_updates = inputs.get("cell_updates")
        column_updates = inputs.get("column_updates")

        # Load workbook (invalid bytes時は新規Workbookにフォールバック)
        try:
            if wb_bytes is not None:
                wb: Workbook = load_workbook(BytesIO(wb_bytes))
            else:
                raise Exception("no-bytes")
        except Exception:
            from openpyxl import Workbook as _WB
            wb = _WB()
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)

        rows_written = 0

        # --- Helper functions -------------------------------------------------
        def _get_ws(_sheet: str) -> Any:
            try:
                if _sheet in wb.sheetnames:
                    return wb[_sheet]
                # 既定の active を Results に改名して使うケースを優先（空ブックの初期シート）
                if len(wb.sheetnames) == 1 and wb.active and wb.active.max_row == 1 and wb.active.max_column == 1:
                    wb.active.title = _sheet
                    return wb[_sheet]
                return wb.create_sheet(_sheet)
            except Exception:
                return wb.create_sheet(_sheet)

        def _as_records(values_obj: Any) -> List[Dict[str, Any]]:
            """Normalize values to list[dict]. Supports list[dict], dict[str, list], pandas.DataFrame."""
            try:
                import pandas as _pd  # type: ignore
            except Exception:  # pragma: no cover - pandas may be unavailable
                _pd = None

            # DataFrame
            if _pd is not None:
                try:
                    if isinstance(values_obj, _pd.DataFrame):
                        return values_obj.to_dict(orient="records")
                except Exception:
                    pass

            # list[dict]
            if isinstance(values_obj, list) and all(isinstance(x, dict) for x in values_obj):
                return values_obj  # type: ignore[return-value]

            # dict[str, list] -> records
            if isinstance(values_obj, dict):
                keys = list(values_obj.keys())
                try:
                    max_len = max(len(values_obj[k]) for k in keys)
                except Exception:
                    # not a dict[str, list]
                    return []
                out: List[Dict[str, Any]] = []
                for i in range(max_len):
                    rec: Dict[str, Any] = {}
                    for k in keys:
                        seq = values_obj.get(k)
                        if isinstance(seq, list) and i < len(seq):
                            rec[k] = seq[i]
                        else:
                            rec[k] = None
                    out.append(rec)
                return out
            return []

        def _col_index(col: Any, default_idx: int) -> int:
            if isinstance(col, int):
                return max(1, col)
            if isinstance(col, str):
                try:
                    return column_index_from_string(col)
                except Exception:
                    return default_idx
            return default_idx

        def _get_by_path(obj: Any, path: str) -> Any:
            try:
                cur = obj
                for raw_seg in str(path).split("."):
                    seg = str(raw_seg).strip()
                    if not seg:
                        return None
                    if isinstance(cur, dict):
                        # case-insensitive key lookup
                        found = None
                        for k in cur.keys():
                            try:
                                if str(k).lower() == seg.lower():
                                    found = k
                                    break
                            except Exception:
                                continue
                        if found is None:
                            return None
                        cur = cur.get(found)
                    else:
                        return None
                return cur
            except Exception:
                return None

        def _deep_find_value(obj: Any, key_name: str, max_depth: int = 3) -> Any:
            """Search nested dict/list for a key (case-insensitive) and return first value found."""
            try:
                if max_depth < 0:
                    return None
                if isinstance(obj, dict):
                    # direct keys
                    for k, v in obj.items():
                        try:
                            if str(k).lower() == str(key_name).lower():
                                return v
                        except Exception:
                            continue
                    # nested search
                    for v in obj.values():
                        found = _deep_find_value(v, key_name, max_depth - 1)
                        if found is not None:
                            return found
                elif isinstance(obj, list):
                    for it in obj:
                        found = _deep_find_value(it, key_name, max_depth - 1)
                        if found is not None:
                            return found
            except Exception:
                return None
            return None

        def _resolve_value(rec: Dict[str, Any], key_or_path: str) -> Any:
            """Resolve value from record by path; fallback to common wrappers and deep search.

            Priority:
              1) direct path (e.g., "row_data.フィールド")
              2) if plain key (no dot): try wrappers 'row_data' and 'results'
              3) deep search by key name in nested dicts/lists (case-insensitive)
            """
            # 1) try as-is (path)
            val = _get_by_path(rec, str(key_or_path))
            if val is not None:
                return val
            # If not a dot path, try common wrapper prefixes
            if "." not in str(key_or_path):
                for wrapper in ("row_data", "results"):
                    if isinstance(rec, dict) and wrapper in rec:
                        val = _get_by_path(rec, f"{wrapper}.{key_or_path}")
                        if val is not None:
                            return val
            # Deep search
            return _deep_find_value(rec, str(key_or_path))

        def _last_segment_name(path: str) -> str:
            try:
                s = str(path)
                if not s:
                    return ""
                seg = s.split(".")[-1].strip()
                return seg
            except Exception:
                return ""

        def _is_excel_col_token(s: Any) -> bool:
            if isinstance(s, int):
                return True
            if not isinstance(s, str):
                return False
            try:
                _ = column_index_from_string(s)
                return True
            except Exception:
                return False

        # --- 1) セル指定更新 ---------------------------------------------------
        if cell_updates is not None:
            ops: List[Dict[str, Any]] = []
            if isinstance(cell_updates, list):
                ops = [x for x in cell_updates if isinstance(x, dict)]
            elif isinstance(cell_updates, dict):
                # {sheet, cells} or {"A1": val, ...}
                if any(k in cell_updates for k in ("sheet", "cells")):
                    ops = [cell_updates]
                else:
                    ops = [{"sheet": sheet_name, "cells": dict(cell_updates)}]

            import re as _re
            addr_re = _re.compile(r"^[A-Za-z]+\d+$")
            for op in ops:
                sh = op.get("sheet") or sheet_name
                ws_local = _get_ws(sh)
                cells: Dict[str, Any] = {}
                raw_cells = op.get("cells")
                if isinstance(raw_cells, dict):
                    cells.update(raw_cells)
                # トップレベルの A1/B2... キーも拾う（互換・安全化）
                for k, v in list(op.items()):
                    try:
                        if isinstance(k, str) and addr_re.match(k):
                            cells.setdefault(k, v)
                    except Exception:
                        continue
                for addr, val in cells.items():
                    try:
                        try:
                            # 文字座標の厳密解釈
                            r_, c_ = coordinate_to_tuple(addr)
                            ws_local.cell(row=r_, column=c_, value=val)
                        except Exception:
                            # フォールバック
                            ws_local[addr].value = val
                        rows_written += 0  # セル更新は行数に加算しない
                    except Exception:
                        continue

        # --- 2) 列指定更新 -----------------------------------------------------
        if column_updates is not None:
            ops2: List[Dict[str, Any]] = []
            if isinstance(column_updates, list):
                ops2 = [x for x in column_updates if isinstance(x, dict)]
            elif isinstance(column_updates, dict):
                ops2 = [column_updates]

            for op in ops2:
                sh = op.get("sheet") or sheet_name
                ws_local = _get_ws(sh)
                start_r = _coerce_int(op.get("start_row"), start_row)
                header_row = _coerce_int(op.get("header_row"), 1)
                # column_map の堅牢化（文字列JSON/None対応）
                col_map_raw = op.get("column_map")
                columns_spec = op.get("columns")
                col_map: Dict[str, Any] = {}
                columns_list: List[Dict[str, str]] = []  # [{"header": str, "path": str}]
                # 1) 新簡易記法: columns（推奨）
                if isinstance(columns_spec, dict):
                    # {header: path, ...}
                    for header, path in columns_spec.items():
                        if isinstance(path, str):
                            columns_list.append({"header": str(header), "path": path})
                elif isinstance(columns_spec, list):
                    for item in columns_spec:
                        if isinstance(item, dict):
                            path = item.get("path") or item.get("field") or item.get("key") or item.get("name")
                            header = item.get("header") or item.get("title") or item.get("name") or _last_segment_name(path or "")
                            if isinstance(path, str) and path:
                                columns_list.append({"header": str(header or _last_segment_name(path)), "path": path})
                        elif isinstance(item, str):
                            columns_list.append({"header": _last_segment_name(item), "path": item})
                # 2) 互換: column_map が JSON 文字列
                if not columns_list and isinstance(col_map_raw, str):
                    try:
                        import json as _json
                        parsed = _json.loads(col_map_raw)
                        if isinstance(parsed, dict):
                            col_map = parsed
                    except Exception:
                        col_map = {}
                # 3) 互換: column_map が dict
                if not columns_list and isinstance(col_map_raw, dict):
                    # 判別: 値が列トークン（"A", 1, など）なら従来記法（path->col）
                    # そうでなければ header->path の簡易記法とみなす
                    if any(_is_excel_col_token(v) for v in col_map_raw.values()):
                        col_map = dict(col_map_raw)
                    else:
                        for header, path in col_map_raw.items():
                            if isinstance(path, str):
                                columns_list.append({"header": str(header), "path": path})
                # values の堅牢化（文字列JSON対応）
                values_obj = op.get("values")
                if isinstance(values_obj, str):
                    try:
                        import json as _json
                        values_obj = _json.loads(values_obj)
                    except Exception:
                        pass
                # Planで与えた values をそのまま解釈し、列マッピング（ドットパス対応）のみ適用
                records = _as_records(values_obj)

                # Optional: clear existing cells in mapped columns before writing
                clear_existing = bool(op.get("clear_existing"))
                if clear_existing:
                    try:
                        last_row = ws_local.max_row or start_r
                        if col_map:
                            cols_iter = list(col_map.values())
                        elif columns_list:
                            cols_iter = list(range(1, len(columns_list) + 1))
                        else:
                            cols_iter = []
                        for col in cols_iter:
                            cidx = _col_index(col, default_idx=1)
                            for rr in range(start_r, last_row + 1):
                                ws_local.cell(row=rr, column=cidx, value=None)
                    except Exception:
                        pass

                # まず新簡易記法（columns_list）があればそれを優先
                if columns_list:
                    # ヘッダー自動書き込み（header_row）
                    write_header = bool(op.get("write_header", True))
                    if write_header:
                        for idx, coldef in enumerate(columns_list, start=1):
                            ws_local.cell(row=header_row, column=idx, value=coldef.get("header"))
                    # 行書き込み
                    r = start_r
                    for rec in records:
                        if not isinstance(rec, dict):
                            continue
                        for idx, coldef in enumerate(columns_list, start=1):
                            path = coldef.get("path") or ""
                            val = _resolve_value(rec, str(path))
                            ws_local.cell(row=r, column=idx, value=val)
                        r += 1
                        rows_written += 1
                else:
                    # 従来記法（path -> 列）
                    # ヘッダー自動書き込み（キーをヘッダと解釈）
                    write_header = bool(op.get("write_header", True))
                    if write_header and isinstance(col_map, dict):
                        for k, col in col_map.items():
                            cidx = _col_index(col, default_idx=1)
                            ws_local.cell(row=header_row, column=cidx, value=k)
                    r = start_r
                    for rec in records:
                        if not isinstance(rec, dict):
                            continue
                        for k, col in col_map.items():
                            cidx = _col_index(col, default_idx=1)
                            # k は path でもヘッダ名でもよい。解決器で補助する。
                            val = _resolve_value(rec, str(k))
                            ws_local.cell(row=r, column=cidx, value=val)
                        r += 1
                        rows_written += 1

        # --- 3) 後方互換: 旧スタイル data + output_config ---------------------
        # items = data.get("items") or []
        # if items:
        #     # Ensure header
        #     if start_row > 1 and ws.max_row < start_row - 1:
        #         ws.cell(row=1, column=1, value=header_map.get("file", "File"))
        #         ws.cell(row=1, column=2, value=header_map.get("count", "Count"))
        #         ws.cell(row=1, column=3, value=header_map.get("sum", "Sum"))

        #     r = start_row
        #     for it in items:
        #         vals = [it.get("file"), it.get("count"), it.get("sum")]
        #         for idx, val in enumerate(vals, start=1):
        #             col = columns[idx - 1] if idx - 1 < len(columns) else get_column_letter(idx)
        #             cidx = _col_index(col, default_idx=idx)
        #             ws.cell(row=r, column=cidx, value=val)
        #         r += 1
        #         rows_written += 1

        # Save back to bytes
        out_buf = BytesIO()
        wb.save(out_buf)
        out_buf.seek(0)
        out_bytes = out_buf.getvalue()

        summary = {"rows_written": rows_written, "sheet": sheet_name, "workbook_name": wb_name}
        # 書き戻ししたブックbytesも返せるように（今後の拡張用）
        # さらに base64 も提供して上位UIで安全に扱えるようにする
        return {
            "write_summary": summary,
            "workbook_updated": {"name": wb_name, "bytes": out_bytes},
            "workbook_b64": base64.b64encode(out_bytes).decode("ascii"),
        }



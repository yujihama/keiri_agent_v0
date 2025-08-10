from __future__ import annotations

from io import BytesIO
from typing import Iterable, List, Tuple


def _read_txt(name: str, data: bytes) -> str:
    try:
        return data.decode("utf-8", errors="ignore")
    except Exception:
        return ""


def _read_md(name: str, data: bytes) -> str:
    return _read_txt(name, data)


def _read_pdf(name: str, data: bytes) -> str:
    try:
        from PyPDF2 import PdfReader

        reader = PdfReader(BytesIO(data))
        texts = []
        for page in reader.pages[:20]:
            try:
                texts.append(page.extract_text() or "")
            except Exception:
                continue
        return "\n".join(texts)
    except Exception:
        return ""


def _read_docx(name: str, data: bytes) -> str:
    try:
        import docx  # python-docx

        doc = docx.Document(BytesIO(data))
        return "\n".join(p.text for p in doc.paragraphs)
    except Exception:
        return ""


def _read_xlsx(name: str, data: bytes) -> str:
    try:
        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
        texts = []
        for ws in wb.worksheets[:2]:
            cnt = 0
            for row in ws.iter_rows(min_row=1, max_row=50, values_only=True):
                vals = [str(v) for v in row if v is not None]
                if vals:
                    texts.append(",".join(vals))
                cnt += 1
                if cnt >= 50:
                    break
        return "\n".join(texts)
    except Exception:
        return ""


def extract_texts(files: Iterable[Tuple[str, bytes]], max_total_chars: int = 100_000) -> List[str]:
    """Extract plain texts from given files.

    files: iterable of (filename, bytes)
    returns: list of text strings per file, trimmed to `max_total_chars` total.
    """

    handlers = {
        ".txt": _read_txt,
        ".md": _read_md,
        ".pdf": _read_pdf,
        ".docx": _read_docx,
        ".xlsx": _read_xlsx,
    }

    out: List[str] = []
    total = 0
    for name, data in files:
        ext = ""
        if "." in name:
            ext = name[name.rfind(".") :].lower()
        handler = handlers.get(ext, _read_txt)
        text = handler(name, data) or ""
        if not text:
            continue
        if total + len(text) > max_total_chars:
            text = text[: max_total_chars - total]
        out.append(text)
        total += len(text)
        if total >= max_total_chars:
            break
    return out



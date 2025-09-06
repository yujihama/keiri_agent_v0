import pandas as pd
import openpyxl
import json

# Excelファイルパス
excel_path = "headless/output/nlp_compare_fixed/text_similarity_compare_two_files/2025-09-05T03-07-32/artifacts/text_compare.xlsx"

print(f"=== Excel Analysis: {excel_path} ===\n")

# openpyxlで詳細確認
try:
    wb = openpyxl.load_workbook(excel_path)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n--- Sheet: {sheet_name} ---")
        print(f"Max row: {ws.max_row}, Max column: {ws.max_column}")
        
        # すべての行を表示（最大20行まで）
        print("\nContent:")
        for row_idx in range(1, min(21, ws.max_row + 1)):
            row_data = []
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                row_data.append(str(cell.value) if cell.value is not None else "")
            print(f"Row {row_idx}: {row_data}")
    
    wb.close()
    
except Exception as e:
    print(f"Error reading Excel with openpyxl: {e}")

# データの詳細を確認
print("\n\n=== Checking intermediate data ===")

# flatten_rows出力を確認
flatten_rows_path = "headless/output/nlp_compare_fixed/text_similarity_compare_two_files/2025-09-05T03-07-32/artifacts/flatten_rows_outputs.json"
try:
    with open(flatten_rows_path, 'r', encoding='utf-8') as f:
        flatten_data = json.load(f)
        print(f"\nflatten_rows output - Total items: {len(flatten_data.get('items', []))}")
        for i, item in enumerate(flatten_data.get('items', [])[:5]):
            print(f"\nItem {i+1}:")
            print(json.dumps(item, ensure_ascii=False, indent=2))
except Exception as e:
    print(f"Error reading flatten_rows output: {e}")

# test_similarity_ab出力を確認
similarity_ab_path = "headless/output/nlp_compare_fixed/text_similarity_compare_two_files/2025-09-05T03-07-32/artifacts/test_similarity_ab_outputs.json"
try:
    with open(similarity_ab_path, 'r', encoding='utf-8') as f:
        similarity_data = json.load(f)
        print(f"\n\ntest_similarity_ab output:")
        print(json.dumps(similarity_data, ensure_ascii=False, indent=2))
except Exception as e:
    print(f"Error reading test_similarity_ab output: {e}")

import pandas as pd
import openpyxl

excel_path = "headless/output/nlp_compare/text_compare_result.xlsx"

print(f"=== Detailed Excel Analysis: {excel_path} ===\n")

# Using openpyxl for detailed inspection
try:
    wb = openpyxl.load_workbook(excel_path)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n--- Sheet: {sheet_name} ---")
        print(f"Max row: {ws.max_row}, Max column: {ws.max_column}")
        
        # Print first 10 rows with all columns
        print("\nContent (first 10 rows):")
        for row_idx in range(1, min(11, ws.max_row + 1)):
            row_data = []
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                row_data.append(str(cell.value) if cell.value is not None else "")
            print(f"Row {row_idx}: {row_data}")
    
    wb.close()
    
except Exception as e:
    print(f"Error reading with openpyxl: {e}")

# Also use pandas for cleaner display
print("\n\n=== Pandas View ===")
try:
    xl = pd.ExcelFile(excel_path)
    
    for sheet_name in xl.sheet_names:
        print(f"\n--- Sheet: {sheet_name} ---")
        df = pd.read_excel(excel_path, sheet_name=sheet_name)
        
        # Show full content if small enough
        if len(df) <= 20:
            print(df.to_string())
        else:
            print(f"First 10 rows:")
            print(df.head(10).to_string())
            print(f"\nLast 5 rows:")
            print(df.tail(5).to_string())
            
except Exception as e:
    print(f"Error reading with pandas: {e}")

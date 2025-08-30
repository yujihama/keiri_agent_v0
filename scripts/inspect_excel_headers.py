from __future__ import annotations

import sys
from pathlib import Path
from typing import Any, Dict, List

# プロジェクトルートを sys.path に追加
PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from core.blocks.base import BlockContext
from core.blocks.processing.excel.read_data import ExcelReadDataBlock


def inspect_with_openpyxl(file_path: Path) -> None:
    """openpyxlで直接Excelファイルの内容を確認"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(file_path), data_only=True)
        print(f"\n=== Direct openpyxl inspection ===")
        print(f"Sheet names: {wb.sheetnames}")

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\nSheet: {sheet_name}")
            print(f"Dimensions: {ws.dimensions}")
            print(f"Max row: {ws.max_row}, Max column: {ws.max_column}")

            # 最初の数行を表示
            rows = list(ws.iter_rows(values_only=True))
            print(f"Total rows from iter_rows: {len(rows)}")

            if rows:
                print("\nFirst row (headers):")
                for i, cell in enumerate(rows[0], 1):
                    print(f"  Col {i}: '{cell}'")

                print("\nFirst 3 data rows:")
                for i, row in enumerate(rows[1:4], 1):
                    if i <= len(rows) - 1:
                        print(f"  Row {i}: {[cell for cell in row[:5]]}...")  # 最初の5カラム

            # Cellオブジェクトで確認
            print("\n=== Cell objects inspection ===")
            for row in range(1, min(4, ws.max_row + 1)):
                row_data = []
                for col in range(1, min(6, ws.max_column + 1)):  # 最初の5列
                    cell = ws.cell(row=row, column=col)
                    row_data.append(f"'{cell.value}'")
                print(f"Row {row}: [{', '.join(row_data)}]")

    except Exception as e:
        print(f"openpyxl error: {e}")
        import traceback
        traceback.print_exc()


def main() -> None:
    # ピボットファイルのヘッダを確認
    pivot_path = Path("tests/data/retirement_data/給与明細一覧表_pivot.xlsx")
    ctx = BlockContext(run_id="inspect")

    print(f"Inspecting: {pivot_path}")
    print(f"File exists: {pivot_path.exists()}")

    if not pivot_path.exists():
        print("File not found!")
        return

    # まず openpyxl で直接確認
    inspect_with_openpyxl(pivot_path)

    # 次にブロック経由で確認
    print("\n=== Block-based inspection ===")

    # 異なる設定で試行
    configs = [
        {
            "name": "header_row=1",
            "config": {
                "header_row": 1,
                "sheets": [{"name": "Sheet1", "header_row": 1}],
                "skip_empty_rows": False,  # 空行も含めて確認
            }
        },
        {
            "name": "header_row=0",
            "config": {
                "header_row": 0,
                "sheets": [{"name": "Sheet1", "header_row": 0}],
                "skip_empty_rows": False,
            }
        },
        {
            "name": "no header_row",
            "config": {
                "sheets": [{"name": "Sheet1"}],
                "skip_empty_rows": False,
            }
        }
    ]

    for config_info in configs:
        print(f"\n--- {config_info['name']} ---")
        try:
            out = ExcelReadDataBlock().run(
                ctx,
                {
                    "workbook": {"path": str(pivot_path)},
                    "mode": "single",
                    "read_config": config_info["config"],
                },
            )

            rows: List[Dict[str, Any]] = out.get("rows") or []
            print(f"Total rows: {len(rows)}")

            if rows:
                first_row = rows[0]
                print("First row keys:")
                for i, key in enumerate(first_row.keys(), 1):
                    print(f"  {i}: '{key}' = {first_row[key]}")
                break  # 成功した設定があればそこで止める
        except Exception as e:
            print(f"Block error: {e}")
            import traceback
            traceback.print_exc()


if __name__ == "__main__":
    main()

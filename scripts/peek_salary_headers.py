from __future__ import annotations

from pathlib import Path
import sys
from typing import Any, Dict, List

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from core.blocks.base import BlockContext
from core.blocks.processing.excel.read_data import ExcelReadDataBlock


def main() -> None:
    path = Path("tests/data/retirement_data/給与明細一覧表.xlsx")
    ctx = BlockContext(run_id="peek")
    # Print sheet names using openpyxl
    try:
        from openpyxl import load_workbook  # type: ignore
        wb = load_workbook(str(path), read_only=True, data_only=True)
        print("sheetnames:", wb.sheetnames)
    except Exception as e:
        print("openpyxl_error:", e)
    out = ExcelReadDataBlock().run(
        ctx,
        {
            "workbook": {"path": str(path)},
            "mode": "single",
            "read_config": {
                "header_row": 1,
                "sheets": [{"name": "Sheet1", "header_row": 1}],
                "skip_empty_rows": True,
            },
        },
    )
    rows: List[Dict[str, Any]] = out.get("rows") or []
    print("rows_count:", len(rows))
    if rows:
        first = rows[0]
        print("first_keys:", list(first.keys()))
        print("first_row:", first)
        # Set of all keys across first 5 rows
        keys = set()
        for r in rows[:5]:
            keys.update(r.keys())
        print("union_keys_first5:", sorted(keys))


if __name__ == "__main__":
    main()



from __future__ import annotations

import json
from pathlib import Path
from typing import Any, List, Dict

from openpyxl import load_workbook


def read_summary_rows(xlsx_path: Path) -> List[Dict[str, Any]]:
    wb = load_workbook(xlsx_path)
    ws = wb["Summary"]
    headers: List[str] = []
    for cell in ws[1]:
        headers.append(str(cell.value or "").strip())
    rows: List[Dict[str, Any]] = []
    for r in ws.iter_rows(min_row=2, max_col=len(headers), values_only=True):
        if not any(c not in (None, "") for c in r):
            continue
        rec = {headers[i]: r[i] for i in range(min(len(headers), len(r)))}
        rows.append(rec)
    return rows


def main() -> None:
    out_dir = Path("headless/output/nlp_compare")
    xlsx = out_dir / "text_compare.xlsx"
    if not xlsx.exists():
        print("EXISTS False")
        return
    rows = read_summary_rows(xlsx)
    print(f"EXISTS True")
    print(f"DATA_ROWS {len(rows)}")
    # プレビュー3件
    print("ROWS_PREVIEW", json.dumps(rows[:3], ensure_ascii=False))


if __name__ == "__main__":
    main()



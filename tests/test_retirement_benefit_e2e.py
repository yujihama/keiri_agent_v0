from __future__ import annotations

import json
from pathlib import Path

import pytest
from dotenv import load_dotenv  # type: ignore
from openpyxl import load_workbook

from core.blocks.registry import BlockRegistry
from core.plan.loader import load_plan
from core.plan.runner import PlanRunner
from core.plan.execution_context import ExecutionContext


def _preferred_path(base_dir: Path, primary: str, fallback: str) -> Path:
    p = base_dir / primary
    if p.exists():
        return p
    old = base_dir / "old" / primary
    if old.exists():
        return old
    # final fallback name
    pf = base_dir / fallback
    if pf.exists():
        return pf
    # old fallback
    of = base_dir / "old" / fallback
    if of.exists():
        return of
    return p


@pytest.mark.e2e
def test_retirement_benefit_q1_2025_e2e(tmp_path: Path):
    # .env からAPIキー等を読み込み（LLM呼び出し用）
    load_dotenv()

    plan_path = Path("designs/retirement_benefit_q1_2025.yaml")
    if not plan_path.exists():
        pytest.skip("Plan file not found; skipping environment-specific E2E")

    reg = BlockRegistry(); reg.load_specs()
    plan = load_plan(plan_path)
    # LLMブロックはテストでは高速化のため fast-path を利用（answer を事前注入）
    for n in plan.graph:
        try:
            if getattr(n, "block", "") == "ai.process_llm":
                inputs = dict(getattr(n, "inputs", {}))
                ev = dict(inputs.get("evidence_data") or {})
                ev["answer"] = "{\"results\": [], \"summary\": {}}"
                inputs["evidence_data"] = ev
                n.inputs = inputs
        except Exception:
            pass

    # 入力ファイル（ヘッドレス実行に切替）
    base_dir = Path("tests/data/retirement_data")
    employees_path = _preferred_path(base_dir, "社員一覧.csv", "給与明細一覧表.csv")
    journal_path = _preferred_path(base_dir, "仕訳データ.csv", "退職給付金勘定元帳.csv")
    workbook_path = _preferred_path(base_dir, "退職給付ワークブック.xlsx", "退職給付ワークブック.xlsx")

    # いずれかが存在しない場合はスキップ
    missing = [p for p in [employees_path, journal_path, workbook_path] if not p.exists()]
    if missing:
        pytest.skip(f"required test data missing: {missing}")

    # ヘッドレス用実行コンテキスト
    exec_ctx = ExecutionContext(
        headless_mode=True,
        output_dir=tmp_path / "out",
        vars_overrides={},
        file_inputs={
            "employees_csv": employees_path,
            "journal_csv": journal_path,
            "workbook": workbook_path,
        },
        ui_mock_responses={
            "ui.interactive_input": {
                "select_quarter": {
                    "collected_data": {"fiscal_year": "2025", "quarter": "Q1"},
                    "approved": True,
                    "metadata": {"submitted": True, "mode": "inquire"},
                }
            }
        },
    )

    # ランナー（HITL不要）
    runs_dir = tmp_path / "runs"
    runner = PlanRunner(registry=reg, runs_dir=runs_dir, default_ui_hitl=False)

    result = runner.run(plan, execution_context=exec_ctx)

    # 出力の基本検証（Excel生成）
    assert isinstance(result, dict)
    wb_out = result.get("wb_final") or result.get("wb") or result.get("workbook_updated")
    assert isinstance(wb_out, dict) and isinstance(wb_out.get("bytes"), (bytes, bytearray)), "workbook bytes missing"

    # Excel内容の検証（ターゲットシート、C2、F2）
    out_file = tmp_path / "retirement_benefit_q1_2025.xlsx"
    out_file.write_bytes(wb_out["bytes"])  # type: ignore[arg-type]
    wb = load_workbook(out_file, data_only=True)
    assert "2025_Q1" in wb.sheetnames
    ws = wb["2025_Q1"]
    assert str(ws["C2"].value) == "2025-06-30"
    f2 = ws["F2"].value
    assert f2 in (None, ""), "F2 must be cleared for Q1"

    # ログが生成されていること
    plan_dir = runs_dir / plan.id
    logs = list(plan_dir.glob("*.jsonl"))
    assert len(logs) >= 1


